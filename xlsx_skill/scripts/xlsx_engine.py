#!/usr/bin/env python3
"""
XLSX Engine - Core utilities for Excel file manipulation
Provides safe loading, saving, reading, editing, recalculation, and error checking
"""

import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
import json
import sys
import argparse
from pathlib import Path


def load_workbook_safe(path):
    """
    Safely load an Excel workbook with error handling
    
    Args:
        path: Path to the .xlsx file
        
    Returns:
        openpyxl Workbook object
        
    Raises:
        FileNotFoundError: If the file doesn't exist
        InvalidFileException: If the file is not a valid Excel file
    """
    path = Path(path)
    
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")
    
    try:
        # data_only=False preserves formulas
        wb = load_workbook(str(path), data_only=False)
        return wb
    except InvalidFileException as e:
        raise InvalidFileException(f"Invalid Excel file format: {path}") from e
    except Exception as e:
        raise Exception(f"Error loading workbook: {e}") from e


def save_workbook_safe(wb, path):
    """
    Safely save an Excel workbook with error handling
    
    Args:
        wb: openpyxl Workbook object
        path: Path to save the .xlsx file
        
    Raises:
        Exception: If saving fails
    """
    try:
        path = Path(path)
        # Create parent directory if it doesn't exist
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
    except Exception as e:
        raise Exception(f"Error saving workbook: {e}") from e


def read_sheet(path, sheet_name=None):
    """
    Extract values and formulas from a specific sheet
    
    Args:
        path: Path to the .xlsx file
        sheet_name: Name of the sheet to read (None = active sheet)
        
    Returns:
        dict: Contains sheet info, cells with coordinates, values, and formulas
    """
    wb = load_workbook_safe(path)
    
    # Get the sheet
    if sheet_name is None:
        sheet = wb.active
        sheet_name = sheet.title
    else:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
        sheet = wb[sheet_name]
    
    result = {
        "sheet_name": sheet_name,
        "cells": []
    }
    
    # Iterate through all cells
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell_data = {
                    "cell": cell.coordinate,
                    "value": cell.value
                }
                
                # Check if it's a formula
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    cell_data["formula"] = cell.value
                
                result["cells"].append(cell_data)
    
    return result


def edit_sheet(path, sheet_name, updates):
    """
    Update cell values or formulas in a sheet
    
    Args:
        path: Path to the .xlsx file
        sheet_name: Name of the sheet to edit
        updates: List of dicts with 'cell' and 'value' keys
                 Example: [{"cell": "A1", "value": 100}, {"cell": "B1", "value": "=A1*2"}]
    
    Returns:
        dict: Status message with updated cells
    """
    wb = load_workbook_safe(path)
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
    
    sheet = wb[sheet_name]
    updated_cells = []
    
    for update in updates:
        cell_ref = update.get("cell")
        value = update.get("value")
        
        if not cell_ref:
            raise ValueError("Each update must have a 'cell' key")
        
        # Update the cell
        sheet[cell_ref] = value
        updated_cells.append(cell_ref)
    
    # Save the workbook
    save_workbook_safe(wb, path)
    
    return {
        "status": "success",
        "updated_cells": updated_cells,
        "count": len(updated_cells)
    }


def recalculate_workbook(path):
    """
    Recalculate all formulas in a workbook using the formulas library
    
    Args:
        path: Path to the .xlsx file
        
    Returns:
        dict: Status and recalculated values
    """
    try:
        import formulas
        
        # Load the workbook model and calculate
        xl_model = formulas.ExcelModel().loads(str(path)).finish()
        results = xl_model.calculate()
        
        # Get the workbook to update cells
        wb = load_workbook_safe(path)
        
        recalculated = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_data = []
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        # Build the reference key
                        cell_ref = f"{sheet_name}!{cell.coordinate}"
                        
                        # Get calculated value from results
                        try:
                            if cell_ref in results:
                                calc_value = results[cell_ref]
                                
                                sheet_data.append({
                                    "cell": cell.coordinate,
                                    "formula": cell.value,
                                    "calculated_value": calc_value
                                })
                            else:
                                sheet_data.append({
                                    "cell": cell.coordinate,
                                    "formula": cell.value,
                                    "note": "Formula not calculated"
                                })
                        except Exception as e:
                            sheet_data.append({
                                "cell": cell.coordinate,
                                "formula": cell.value,
                                "error": f"Calculation error: {str(e)}"
                            })
            
            if sheet_data:
                recalculated[sheet_name] = sheet_data
        
        return {
            "status": "success",
            "recalculated": recalculated,
            "note": "Formulas have been recalculated. Values shown are computed results."
        }
        
    except ImportError:
        return {
            "status": "error",
            "message": "formulas library not installed. Install with: pip install 'formulas[excel]'",
            "type": "ImportError"
        }
    except FileNotFoundError as e:
        return {
            "status": "error",
            "message": str(e),
            "type": "FileNotFoundError"
        }
    except Exception as e:
        return {
            "status": "error",
            "message": str(e),
            "type": type(e).__name__,
            "note": "Some Excel functions may not be supported by the formulas library. Supported functions include basic arithmetic, SUM, AVERAGE, IF, VLOOKUP, and more. Complex or VBA-based functions are not supported."
        }


def check_errors(path):
    """
    Scan for Excel error codes and report them
    
    Args:
        path: Path to the .xlsx file
        
    Returns:
        dict: List of errors found with cell coordinates and error types
    """
    wb = load_workbook_safe(path)
    
    # Common Excel error codes
    error_codes = ['#REF!', '#DIV/0!', '#NAME?', '#VALUE!', '#N/A', '#NULL!', '#NUM!']
    
    errors = []
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value in error_codes:
                    errors.append({
                        "sheet": sheet_name,
                        "cell": cell.coordinate,
                        "error": cell.value,
                        "formula": cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else None
                    })
    
    return {
        "errors_found": len(errors),
        "errors": errors
    }


def main():
    """CLI interface for xlsx_engine"""
    parser = argparse.ArgumentParser(description='Excel file manipulation engine')
    subparsers = parser.add_subparsers(dest='command', help='Commands')
    
    # Read command
    read_parser = subparsers.add_parser('read', help='Read sheet data')
    read_parser.add_argument('--path', required=True, help='Path to Excel file')
    read_parser.add_argument('--sheet', default=None, help='Sheet name (default: active sheet)')
    
    # Edit command
    edit_parser = subparsers.add_parser('edit', help='Edit sheet data')
    edit_parser.add_argument('--path', required=True, help='Path to Excel file')
    edit_parser.add_argument('--sheet', required=True, help='Sheet name')
    edit_parser.add_argument('--updates', required=True, help='JSON string of updates')
    
    # Check errors command
    errors_parser = subparsers.add_parser('check_errors', help='Check for errors')
    errors_parser.add_argument('--path', required=True, help='Path to Excel file')
    
    # Recalculate command
    recalc_parser = subparsers.add_parser('recalculate', help='Recalculate formulas')
    recalc_parser.add_argument('--path', required=True, help='Path to Excel file')
    
    args = parser.parse_args()
    
    try:
        if args.command == 'read':
            result = read_sheet(args.path, args.sheet)
            print(json.dumps(result, indent=2))
            
        elif args.command == 'edit':
            updates = json.loads(args.updates)
            result = edit_sheet(args.path, args.sheet, updates)
            print(json.dumps(result, indent=2))
            
        elif args.command == 'check_errors':
            result = check_errors(args.path)
            print(json.dumps(result, indent=2))
            
        elif args.command == 'recalculate':
            result = recalculate_workbook(args.path)
            print(json.dumps(result, indent=2))
            
        else:
            parser.print_help()
            sys.exit(1)
            
    except Exception as e:
        error_result = {
            "status": "error",
            "message": str(e),
            "type": type(e).__name__
        }
        print(json.dumps(error_result, indent=2))
        sys.exit(1)


if __name__ == '__main__':
    main()
